"""On-demand reports (PDF + Excel) for the admin and branch Reports pages.

Design intent:
- One reusable pattern instead of a bespoke page per report: pick a report
  TYPE, pick a time window (Recent N / Date range / All time), pick that
  type's own extra filters, then download as PDF or Excel. Admin.py and
  branch.py both call get_report() + render_report_pdf()/render_report_excel()
  — the branch side just always passes branch_scope=<their own branch_id>,
  which quietly removes the Branch column and ignores any branch_id filter
  the querystring might contain, so a branch user can never pull another
  branch's report by editing the URL.
- Every report is capped at MAX_ROWS rows even on "All time" — see the
  `truncated` flag in the returned dict, which both templates surface to
  the user rather than silently dropping rows.
- Numbers/dates are kept as native Python types (Decimal, date, datetime)
  in the row dicts for as long as possible, and only formatted to strings
  right before they're placed on the page — Excel gets real numbers/dates
  with number formats, PDF gets formatted strings.
"""
import datetime
import io
import os

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

from db import query
from utils import PARTNER_TYPES, PAYMENT_METHODS, PRODUCT_UNITS, SALE_TYPES

# Built-in PDF fonts (Helvetica etc.) only cover Latin-1 and have no glyph
# for the ₱ (Philippine peso) sign — it silently renders as a black "tofu"
# box instead of erroring, which is easy to miss until someone opens the
# PDF. DejaVu Sans does have that glyph, so every style below uses it
# instead. Bundled under fonts/ so this works the same on any machine
# this app runs on, regardless of what fonts happen to be installed
# system-wide.
_FONTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
pdfmetrics.registerFont(
    TTFont("DejaVuSans", os.path.join(_FONTS_DIR, "DejaVuSans.ttf")))
pdfmetrics.registerFont(
    TTFont("DejaVuSans-Bold", os.path.join(_FONTS_DIR, "DejaVuSans-Bold.ttf")))

MAX_ROWS = 1000
RECENT_CHOICES = (20, 50, 100, 200)
STATUS_CHOICES = ("Pending", "In Transit", "Fulfilled", "Rejected")
MOVEMENT_TYPE_CHOICES = ("PRODUCTION", "DISPATCH",
                         "RECEIPT", "SALE", "REFILL", "ADJUSTMENT", "DAMAGE")
VARIANT_CHOICES = ("Male", "Female", "Unisex")
ROLE_CHOICES = ("Admin", "Branch")
UNIT_CHOICES = PRODUCT_UNITS
SALE_TYPE_CHOICES = SALE_TYPES
PAYMENT_METHOD_CHOICES = PAYMENT_METHODS
PARTNER_TYPE_CHOICES = PARTNER_TYPES
# Mirrors packages.partner_scope's ENUM ('Both' plus each PARTNER_TYPES
# value) — kept separate from PARTNER_TYPE_CHOICES since a package can
# also be scoped to "Both", which a partner itself never is.
PACKAGE_SCOPE_CHOICES = ("Both",) + PARTNER_TYPES
# packages.is_active as it reads on screen (Packages page's Status
# column) rather than the raw boolean.
PACKAGE_STATUS_CHOICES = ("Active", "Retired")
# Mirrors partner_inquiries.status's ENUM — same pipeline shown on the
# Partner Inquiries page and _macros.html's inquiry_status_badge.
INQUIRY_STATUS_CHOICES = ("New", "Contacted", "Follow-up",
                          "On Hold", "Closed", "Declined")

INK = colors.HexColor("#12141A")
INK_FAINT = colors.HexColor("#5B6272")
ACCENT = colors.HexColor("#2E5AF0")
ACCENT_INK = colors.HexColor("#1D3BC4")
ACCENT_SOFT = colors.HexColor("#E7ECFE")
BORDER = colors.HexColor("#E5E8EF")
ROW_ALT = colors.HexColor("#FAFBFD")

# Brand red — the "Angel" half of the wordmark, and the same red the
# badge system uses for Female/Rejected/danger states. Kept next to
# ACCENT (blue, "Heaven") so report headers can render the same
# two-tone brand mark used everywhere else (sidebar, login, receipts)
# instead of printing the whole name in one flat color.
BRAND_RED = colors.HexColor("#E23A48")
BRAND_RED_INK = colors.HexColor("#B31E30")

# Hex pairs (background, text) for each badge "style" — a direct port
# of the badge-* classes in style.css (light-theme values), so a
# Status/Type/Variant column in a generated report is colored exactly
# like the matching badge the person already sees on screen.
BADGE_STYLES = {
    "pending":   ("#FBF0D9", "#C9820B"),  # --warning-soft / --warning
    "transit":   ("#E7ECFE", "#1D3BC4"),  # --blue-soft / --blue-ink
    "fulfilled": ("#E1F5EC", "#17975E"),  # --success-soft / --success
    "rejected":  ("#FCE7EA", "#B31E30"),  # --red-soft / --red-ink
    "active":    ("#E1F5EC", "#17975E"),
    "inactive":  ("#F4F5F8", "#97A0AF"),  # --bg / --ink-faint
    "male":      ("#E7ECFE", "#1D3BC4"),
    "female":    ("#FCE7EA", "#B31E30"),
    "unisex":    ("#EFEAFE", "#7C5CFA"),  # --plum-soft / --plum
}

# Maps a column's *semantic kind* (not its literal value) to the
# badge-style key for each value it can hold. Mirrors the macro logic
# in _macros.html (status_badge / movement_badge / sale_type_badge /
# payment_method_badge / variant_badge) line for line, so report
# coloring can never drift out of sync with what the web UI shows for
# that same value.
BADGE_KIND_MAPS = {
    "status": {"Pending": "pending", "In Transit": "transit", "Fulfilled": "fulfilled", "Rejected": "rejected"},
    "movement_type": {
        "PRODUCTION": "fulfilled", "DISPATCH": "transit", "RECEIPT": "fulfilled",
        "SALE": "unisex", "REFILL": "female", "ADJUSTMENT": "pending", "DAMAGE": "rejected",
    },
    "sale_type": {"Sale": "fulfilled", "Refill": "transit"},
    "payment_method": {"Cash": "active", "Salary Deduction": "pending"},
    "variant": {"Male": "male", "Female": "female", "Unisex": "unisex"},
    "active_status": {"Active": "active", "Discontinued": "inactive", "Deactivated": "inactive"},
    "role": {"Admin": "unisex", "Branch": "transit"},
    # Mirrors partners.html / partner_inquiries.html's inline badge:
    # badge-transit for Distributor, badge-unisex for Reseller.
    "partner_type": {"Distributor": "transit", "Reseller": "unisex"},
    # Mirrors packages.html's "For" column.
    "package_scope": {"Both": "pending", "Distributor": "transit", "Reseller": "unisex"},
    # Mirrors packages.html's Status column (Active/Retired badge).
    "package_status": {"Active": "active", "Retired": "inactive"},
    # Mirrors _macros.html's inquiry_status_badge line for line.
    "inquiry_status": {
        "New": "pending", "Contacted": "transit", "Follow-up": "unisex",
        "On Hold": "inactive", "Closed": "fulfilled", "Declined": "rejected",
    },
}


def _badge_colors(kind, value):
    """Return (bg_hex, text_hex) for a badge-kind column's value, or
    None to fall back to plain text (unrecognized kind/value)."""
    style_key = BADGE_KIND_MAPS.get(kind, {}).get(value)
    return BADGE_STYLES.get(style_key) if style_key else None


XL_HEADER_FILL = PatternFill("solid", fgColor="2E5AF0")
XL_TITLE_FILL = PatternFill("solid", fgColor="E7ECFE")
XL_BORDER = Border(*(Side(style="thin", color="E5E8EF"),) * 4)
XL_MONEY_FMT = '"₱"#,##0.00'
XL_DATE_FMT = "mmm dd, yyyy hh:mm AM/PM"

# ---------------------------------------------------------------- registry
# admin / branch: whether that role can generate this report at all.
# windowed: whether a time-window (Recent/Range/All) control applies.
REPORT_TYPES = {
    "products":        {"label": "Products",        "admin": True, "branch": False, "windowed": False},
    "production_log":  {"label": "Production Log",  "admin": True, "branch": False, "windowed": True},
    "branch_stock":    {"label": "Branch Stock",     "admin": True, "branch": True,  "windowed": False,
                        "branch_label": "My Inventory"},
    "stock_requests":  {"label": "Stock Requests",   "admin": True, "branch": True,  "windowed": True},
    "inventory_log":   {"label": "Inventory Log",    "admin": True, "branch": True,  "windowed": True},
    "sales_history":   {"label": "Sales History",    "admin": True, "branch": True,  "windowed": True},
    "employee_purchases": {"label": "Employee Purchases (Salary Deduction)", "admin": True, "branch": True,
                           "windowed": True, "branch_label": "Employee Purchases (Salary Deduction)"},
    "accounts":        {"label": "Accounts",         "admin": True, "branch": False, "windowed": False},
    # Partners & Distribution — admin-only (branch accounts never see
    # this section of the app at all, same as Accounts above).
    # Partners/Packages are catalog-style snapshots (not windowed),
    # same reasoning as Products/Accounts. Partner Inquiries is a dated
    # pipeline of leads, so it gets the Recent/Range/All time window.
    "partners":        {"label": "Partners",         "admin": True, "branch": False, "windowed": False},
    "packages":        {"label": "Packages",         "admin": True, "branch": False, "windowed": False},
    "partner_inquiries": {"label": "Partner Inquiries", "admin": True, "branch": False, "windowed": True},
}


# ---------------------------------------------------------------- filter parsing
def parse_report_filters(args):
    """Sanitize raw querystring args into a plain dict of known-safe values.

    Never trusts a raw value into SQL directly — every field is either
    checked against a fixed allow-list, cast to int, or parsed as a date.
    Anything invalid or missing quietly falls back to a safe default
    rather than erroring, since this only ever affects what's *filtered*,
    not whether the request is allowed.
    """
    def _choice(name, choices, default):
        v = args.get(name, default)
        return v if v in choices else default

    def _date(name):
        raw = (args.get(name) or "").strip()
        try:
            return datetime.datetime.strptime(raw, "%Y-%m-%d").date().isoformat()
        except ValueError:
            return None

    mode = _choice("mode", ("recent", "range", "all"), "recent")
    try:
        recent_n = int(args.get("recent_n", 20))
    except (TypeError, ValueError):
        recent_n = 20
    if recent_n not in RECENT_CHOICES:
        recent_n = 20

    branch_id_raw = (args.get("branch_id") or "all").strip()
    branch_id = int(branch_id_raw) if branch_id_raw.isdigit() else "all"

    return {
        "mode": mode,
        "recent_n": recent_n,
        "date_from": _date("date_from"),
        "date_to": _date("date_to"),
        "branch_id": branch_id,
        "status": _choice("status", STATUS_CHOICES, "all"),
        "movement_type": _choice("movement_type", MOVEMENT_TYPE_CHOICES, "all"),
        "variant": _choice("variant", VARIANT_CHOICES, "all"),
        "unit": _choice("unit", UNIT_CHOICES, "all"),
        "sale_type": _choice("sale_type", SALE_TYPE_CHOICES, "all"),
        "payment_method": _choice("payment_method", PAYMENT_METHOD_CHOICES, "all"),
        "role": _choice("role", ROLE_CHOICES, "all"),
        "account_status": _choice("account_status", ("active", "inactive"), "all"),
        "low_stock_only": args.get("low_stock_only") == "1",
        "partner_type": _choice("partner_type", PARTNER_TYPE_CHOICES, "all"),
        "package_scope": _choice("package_scope", PACKAGE_SCOPE_CHOICES, "all"),
        "package_status": _choice("package_status", PACKAGE_STATUS_CHOICES, "all"),
        "inquiry_status": _choice("inquiry_status", INQUIRY_STATUS_CHOICES, "all"),
        "search": (args.get("search") or "").strip()[:100],
    }


def _time_window(date_col, filters, params):
    """Append a time-window WHERE fragment for date_col and return
    (where_fragment, order_by_sql, row_limit, is_capped_all_time).

    params is mutated in place (range mode appends its bound(s)).
    """
    mode = filters["mode"]
    if mode == "range":
        frag = ""
        if filters["date_from"]:
            frag += f" AND {date_col} >= %s"
            params.append(f"{filters['date_from']} 00:00:00")
        if filters["date_to"]:
            frag += f" AND {date_col} <= %s"
            params.append(f"{filters['date_to']} 23:59:59")
        # Range mode is capped at MAX_ROWS just like "all time" — it was
        # previously hardcoded to False here, so a date range matching
        # more than MAX_ROWS rows would silently drop the excess with no
        # "capped, narrow your filters" note anywhere in the report.
        return frag, f"ORDER BY {date_col} ASC", MAX_ROWS, True
    if mode == "all":
        return "", f"ORDER BY {date_col} DESC", MAX_ROWS, True
    return "", f"ORDER BY {date_col} DESC", filters["recent_n"], False


def _window_note(filters, truncated):
    mode = filters["mode"]
    if mode == "recent":
        note = f"Most recent {filters['recent_n']} entries"
    elif mode == "range":
        frm = filters["date_from"] or "the beginning"
        to = filters["date_to"] or "today"
        note = f"{frm} through {to}"
    else:
        note = "All time"
    if truncated:
        note += f" (capped at the first {MAX_ROWS:,} rows — narrow the filters for a complete report)"
    return note


def _branch_name(branch_id):
    if branch_id in (None, "all"):
        return None
    row = query("SELECT branch_name FROM branches WHERE branch_id = %s",
                (branch_id,), fetchone=True)
    return row["branch_name"] if row else None


# ---------------------------------------------------------------- per-type builders
def _report_products(filters, branch_scope):
    # NOTE: products.is_active was dropped from the schema (products are
    # now edited in place from the admin Products page instead of being
    # discontinued/reactivated — see schema.sql's migration block), so
    # there is no active/discontinued status left to filter or show here.
    where, params = "", []
    if filters["variant"] != "all":
        where += " AND p.variant = %s"
        params.append(filters["variant"])
    if filters["unit"] != "all":
        where += " AND p.unit = %s"
        params.append(filters["unit"])
    if filters["search"]:
        where += " AND (p.item_name LIKE %s OR p.sku LIKE %s)"
        like = f"%{filters['search']}%"
        params += [like, like]

    rows = query(
        f"""SELECT p.sku, p.item_name, p.variant, p.unit, p.price,
                   COALESCE(SUM(bi.stock_qty), 0) AS total_stock
            FROM products p LEFT JOIN branch_inventory bi ON p.sku = bi.sku
            WHERE 1=1 {where}
            GROUP BY p.sku ORDER BY p.item_name LIMIT {MAX_ROWS}""",
        tuple(params),
    )
    for r in rows:
        r["price"] = float(r["price"])

    columns = [
        ("sku", "SKU", "str"), ("item_name", "Item",
                                "str"), ("variant", "Variant", "badge:variant"),
        ("unit", "Unit", "str"), ("price", "HQ Price", "money"),
        ("total_stock", "Total Stock (all branches)", "int"),
    ]
    return columns, rows, len(rows) == MAX_ROWS, "Snapshot as of now"


def _report_production_log(filters, branch_scope):
    where, params = "", []
    if filters["unit"] != "all":
        where += " AND p.unit = %s"
        params.append(filters["unit"])
    if filters["search"]:
        where += " AND (p.item_name LIKE %s OR p.sku LIKE %s OR pl.batch_code LIKE %s)"
        like = f"%{filters['search']}%"
        params += [like, like, like]
    time_where, order, limit_n, truncated = _time_window(
        "pl.produced_at", filters, params)
    where += time_where

    rows = query(
        f"""SELECT pl.produced_at, p.sku, p.item_name, p.unit, pl.batch_code, pl.qty_produced
            FROM production_logs pl JOIN products p ON pl.sku = p.sku
            WHERE 1=1 {where} {order} LIMIT {limit_n}""",
        tuple(params),
    )
    for r in rows:
        r["batch_code"] = r["batch_code"] or "—"

    columns = [
        ("produced_at", "Produced", "datetime"), ("sku",
                                                  "SKU", "str"), ("item_name", "Item", "str"),
        ("unit", "Unit", "str"), ("batch_code", "Batch",
                                  "str"), ("qty_produced", "Qty Produced", "int"),
    ]
    truncated = truncated and len(rows) == MAX_ROWS
    return columns, rows, truncated, _window_note(filters, truncated)


def _report_branch_stock(filters, branch_scope):
    where, params = "", []
    if branch_scope is not None:
        where += " AND b.branch_id = %s"
        params.append(branch_scope)
    elif filters["branch_id"] != "all":
        where += " AND b.branch_id = %s"
        params.append(filters["branch_id"])
    if filters["variant"] != "all":
        where += " AND p.variant = %s"
        params.append(filters["variant"])
    if filters["unit"] != "all":
        where += " AND p.unit = %s"
        params.append(filters["unit"])
    if filters["search"]:
        where += " AND (p.item_name LIKE %s OR p.sku LIKE %s)"
        like = f"%{filters['search']}%"
        params += [like, like]
    if filters["low_stock_only"]:
        where += " AND bi.stock_qty <= bi.reorder_level"

    # No more per-branch price override — every branch sells at
    # products.price, so this is just stock levels per branch now.
    # (products.is_active no longer exists — see schema.sql's migration
    # block — so there's nothing to filter out here anymore.)
    rows = query(
        f"""SELECT b.branch_name, p.sku, p.item_name, p.variant, p.unit, p.price AS hq_price,
                   bi.stock_qty, bi.reorder_level
            FROM branch_inventory bi
            JOIN branches b ON bi.branch_id = b.branch_id
            JOIN products p ON bi.sku = p.sku
            WHERE b.is_hq = FALSE {where}
            ORDER BY b.branch_name, p.item_name LIMIT {MAX_ROWS}""",
        tuple(params),
    )
    for r in rows:
        r["hq_price"] = float(r["hq_price"])

    columns = [("branch_name", "Branch", "str")
               ] if branch_scope is None else []
    columns += [
        ("sku", "SKU", "str"), ("item_name", "Item",
                                "str"), ("variant", "Variant", "badge:variant"),
        ("unit", "Unit", "str"), ("hq_price", "HQ Price", "money"),
        ("stock_qty", "Stock Qty", "int"), ("reorder_level", "Reorder Level", "int"),
    ]
    return columns, rows, len(rows) == MAX_ROWS, "Snapshot as of now"


def _report_stock_requests(filters, branch_scope):
    """One row per product on a delivery.

    A stock request is now a delivery *header* (stock_requests) that can
    carry several products, each its own line in stock_request_items —
    sku/requested_qty/dispatched_qty/received_qty/damaged_qty all live on
    the item row now, not on the request itself (see schema.sql's
    migration block and receipts.py). So this joins through
    stock_request_items rather than reading those columns off sr
    directly, and surfaces delivery_number (the human-facing identifier
    used everywhere else in the app) alongside them.
    """
    where, params = "", []
    if branch_scope is not None:
        where += " AND sr.branch_id = %s"
        params.append(branch_scope)
    elif filters["branch_id"] != "all":
        where += " AND sr.branch_id = %s"
        params.append(filters["branch_id"])
    if filters["status"] != "all":
        where += " AND sr.status = %s"
        params.append(filters["status"])
    if filters["search"]:
        where += " AND (p.item_name LIKE %s OR p.sku LIKE %s OR sr.delivery_number LIKE %s)"
        like = f"%{filters['search']}%"
        params += [like, like, like]
    time_where, order, limit_n, truncated = _time_window(
        "sr.requested_at", filters, params)
    where += time_where

    rows = query(
        f"""SELECT sr.requested_at, sr.delivery_number, b.branch_name, p.item_name, p.sku,
                   sri.requested_qty, sri.dispatched_qty, sri.received_qty, sri.damaged_qty, sr.status
            FROM stock_request_items sri
            JOIN stock_requests sr ON sri.request_id = sr.request_id
            JOIN branches b ON sr.branch_id = b.branch_id
            JOIN products p ON sri.sku = p.sku
            WHERE 1=1 {where} {order} LIMIT {limit_n}""",
        tuple(params),
    )
    for r in rows:
        for k in ("dispatched_qty", "received_qty", "damaged_qty"):
            r[k] = r[k] or 0

    columns = [] if branch_scope is not None else [
        ("branch_name", "Branch", "str")]
    columns = [
        ("requested_at", "Requested", "datetime"),
        ("delivery_number", "Delivery #", "str"),
    ] + columns + [
        ("item_name", "Item", "str"), ("sku", "SKU",
                                       "str"), ("requested_qty", "Requested Qty", "int"),
        ("dispatched_qty", "Dispatched Qty",
         "int"), ("received_qty", "Received Qty", "int"),
        ("damaged_qty", "Damaged Qty", "int"), ("status", "Status", "badge:status"),
    ]
    truncated = truncated and len(rows) == MAX_ROWS
    return columns, rows, truncated, _window_note(filters, truncated)


def _report_inventory_log(filters, branch_scope):
    where, params = "", []
    if branch_scope is not None:
        where += " AND sml.branch_id = %s"
        params.append(branch_scope)
    elif filters["branch_id"] != "all":
        where += " AND sml.branch_id = %s"
        params.append(filters["branch_id"])
    if filters["movement_type"] != "all":
        where += " AND sml.movement_type = %s"
        params.append(filters["movement_type"])
    if filters["search"]:
        where += " AND (p.item_name LIKE %s OR p.sku LIKE %s OR sml.notes LIKE %s)"
        like = f"%{filters['search']}%"
        params += [like, like, like]
    time_where, order, limit_n, truncated = _time_window(
        "sml.created_at", filters, params)
    where += time_where

    rows = query(
        f"""SELECT sml.created_at, b.branch_name, p.item_name, p.sku,
                   sml.movement_type, sml.change_qty, sml.notes
            FROM stock_movement_logs sml
            JOIN branches b ON sml.branch_id = b.branch_id
            JOIN products p ON sml.sku = p.sku
            WHERE 1=1 {where} {order} LIMIT {limit_n}""",
        tuple(params),
    )
    for r in rows:
        r["notes"] = r["notes"] or "—"

    columns = [] if branch_scope is not None else [
        ("branch_name", "Branch", "str")]
    columns = [("created_at", "When", "datetime")] + columns + [
        ("item_name", "Item", "str"), ("sku", "SKU",
                                       "str"), ("movement_type", "Type", "badge:movement_type"),
        ("change_qty", "Change", "int"), ("notes", "Notes", "str"),
    ]
    truncated = truncated and len(rows) == MAX_ROWS
    return columns, rows, truncated, _window_note(filters, truncated)


def _report_sales_history(filters, branch_scope):
    where, params = "", []
    if branch_scope is not None:
        where += " AND s.branch_id = %s"
        params.append(branch_scope)
    elif filters["branch_id"] != "all":
        where += " AND s.branch_id = %s"
        params.append(filters["branch_id"])
    if filters["variant"] != "all":
        where += " AND p.variant = %s"
        params.append(filters["variant"])
    if filters["unit"] != "all":
        where += " AND p.unit = %s"
        params.append(filters["unit"])
    if filters["sale_type"] != "all":
        where += " AND s.sale_type = %s"
        params.append(filters["sale_type"])
    if filters["payment_method"] != "all":
        where += " AND s.payment_method = %s"
        params.append(filters["payment_method"])
    if filters["search"]:
        where += " AND (p.item_name LIKE %s OR p.sku LIKE %s)"
        like = f"%{filters['search']}%"
        params += [like, like]
    time_where, order, limit_n, truncated = _time_window(
        "s.sold_at", filters, params)
    where += time_where

    rows = query(
        f"""SELECT s.sold_at, b.branch_name, p.item_name, p.sku, p.variant, p.unit,
                   s.qty_sold, s.unit_price, (s.qty_sold * s.unit_price) AS line_total,
                   s.sale_type, s.payment_method, COALESCE(s.buyer_name, bu.username) AS buyer_username
            FROM sales s
            JOIN branches b ON s.branch_id = b.branch_id
            JOIN products p ON s.sku = p.sku
            LEFT JOIN users bu ON s.buyer_user_id = bu.user_id
            WHERE 1=1 {where} {order} LIMIT {limit_n}""",
        tuple(params),
    )
    for r in rows:
        r["unit_price"] = float(r["unit_price"])
        r["line_total"] = float(r["line_total"])
        r["buyer_username"] = r["buyer_username"] or "—"

    columns = [] if branch_scope is not None else [
        ("branch_name", "Branch", "str")]
    columns = [("sold_at", "Sold", "datetime")] + columns + [
        ("item_name", "Item", "str"), ("sku", "SKU",
                                       "str"), ("variant", "Variant", "badge:variant"),
        ("unit", "Unit", "str"), ("sale_type",
                                  "Type", "badge:sale_type"), ("qty_sold", "Qty", "int"),
        ("unit_price", "Unit Price", "money"), ("line_total", "Total", "money"),
        ("payment_method", "Payment", "badge:payment_method"), ("buyer_username",
                                                                "Employee (if salary deduction)", "str"),
    ]
    truncated = truncated and len(rows) == MAX_ROWS
    return columns, rows, truncated, _window_note(filters, truncated)


def _report_employee_purchases(filters, branch_scope):
    """Sales paid for via payroll deduction rather than cash — i.e. an
    employee took product for themselves and the cost comes out of their
    salary. Same shape as sales_history but always scoped to
    payment_method = 'Salary Deduction', so HQ/payroll can pull exactly
    what needs to be deducted from each employee for a given period.
    """
    where, params = "", ["Salary Deduction"]
    if branch_scope is not None:
        where += " AND s.branch_id = %s"
        params.append(branch_scope)
    elif filters["branch_id"] != "all":
        where += " AND s.branch_id = %s"
        params.append(filters["branch_id"])
    if filters["sale_type"] != "all":
        where += " AND s.sale_type = %s"
        params.append(filters["sale_type"])
    if filters["search"]:
        where += " AND (p.item_name LIKE %s OR p.sku LIKE %s OR COALESCE(s.buyer_name, bu.username) LIKE %s)"
        like = f"%{filters['search']}%"
        params += [like, like, like]
    time_where, order, limit_n, truncated = _time_window(
        "s.sold_at", filters, params)
    where += time_where

    rows = query(
        f"""SELECT s.sold_at, b.branch_name, p.item_name, p.sku, s.sale_type,
                   s.qty_sold, s.unit_price, (s.qty_sold * s.unit_price) AS line_total,
                   COALESCE(s.buyer_name, bu.username, '(unspecified)') AS buyer_username
            FROM sales s
            JOIN branches b ON s.branch_id = b.branch_id
            JOIN products p ON s.sku = p.sku
            LEFT JOIN users bu ON s.buyer_user_id = bu.user_id
            WHERE s.payment_method = %s {where} {order} LIMIT {limit_n}""",
        tuple(params),
    )
    for r in rows:
        r["unit_price"] = float(r["unit_price"])
        r["line_total"] = float(r["line_total"])

    columns = [] if branch_scope is not None else [
        ("branch_name", "Branch", "str")]
    columns = [("sold_at", "Date", "datetime"), ("buyer_username", "Employee", "str")] + columns + [
        ("item_name", "Item", "str"), ("sku", "SKU",
                                       "str"), ("sale_type", "Type", "badge:sale_type"),
        ("qty_sold", "Qty", "int"), ("unit_price", "Unit Price", "money"),
        ("line_total", "Amount to Deduct", "money"),
    ]
    truncated = truncated and len(rows) == MAX_ROWS
    return columns, rows, truncated, _window_note(filters, truncated)


def _report_accounts(filters, branch_scope):
    where, params = "", []
    if filters["role"] != "all":
        where += " AND u.role = %s"
        params.append(filters["role"])
    if filters["account_status"] != "all":
        where += " AND u.is_active = %s"
        params.append(filters["account_status"] == "active")
    if filters["search"]:
        where += " AND u.username LIKE %s"
        params.append(f"%{filters['search']}%")

    rows = query(
        f"""SELECT u.username, u.role, b.branch_name, u.is_active, u.created_at
            FROM users u LEFT JOIN branches b ON u.branch_id = b.branch_id
            WHERE 1=1 {where} ORDER BY u.role, b.branch_name, u.username LIMIT {MAX_ROWS}""",
        tuple(params),
    )
    for r in rows:
        r["branch_name"] = r["branch_name"] or "— HQ —"
        r["is_active"] = "Active" if r["is_active"] else "Deactivated"

    columns = [
        ("username", "Username", "str"), ("role", "Role",
                                          "badge:role"), ("branch_name", "Branch", "str"),
        ("is_active", "Status", "badge:active_status"), ("created_at",
                                                         "Created", "datetime"),
    ]
    return columns, rows, len(rows) == MAX_ROWS, "Snapshot as of now"


def _report_partners(filters, branch_scope):
    """One row per partner. total_sales/closed_order_count are computed
    the same way the Partners page computes "package sales" — only
    inquiries an admin has marked Closed count, same rule the Partners
    and Dashboard pages both follow (see partners.html's footnote).

    Contact Person/Phone/Email collapse into one Contact column, and
    Total/Closed Inquiries collapse into one Inquiries column — same
    shape the Partners page itself already shows (see the "Contact" and
    "Package sales" cells in partners.html), so the report isn't more
    spread out than the screen it's summarizing. Address is dropped
    entirely: it's an optional field on the partner-portal form and is
    almost always blank in practice, so keeping its own column mostly
    just added width for a column that read "—" on nearly every row.
    """
    where, params = "", []
    if filters["partner_type"] != "all":
        where += " AND p.partner_type = %s"
        params.append(filters["partner_type"])
    if filters["search"]:
        where += " AND (p.partner_name LIKE %s OR p.contact_person LIKE %s OR p.phone LIKE %s OR p.email LIKE %s)"
        like = f"%{filters['search']}%"
        params += [like, like, like, like]

    rows = query(
        f"""SELECT p.partner_name, p.partner_type, p.contact_person, p.phone, p.email,
                   p.inquiry_count, p.last_inquiry_at, p.created_at,
                   COALESCE(SUM(CASE WHEN pi.status = 'Closed' THEN pi.order_amount ELSE 0 END), 0) AS total_sales,
                   COUNT(CASE WHEN pi.status = 'Closed' THEN 1 END) AS closed_order_count
            FROM partners p
            LEFT JOIN partner_inquiries pi ON pi.partner_id = p.partner_id
            WHERE 1=1 {where}
            GROUP BY p.partner_id
            ORDER BY p.partner_name LIMIT {MAX_ROWS}""",
        tuple(params),
    )
    for r in rows:
        r["total_sales"] = float(r["total_sales"])
        r["contact"] = " · ".join(
            p for p in (r.pop("contact_person"), r.pop("phone"), r.pop("email")) if p
        ) or "—"
        r["inquiries_summary"] = f"{r['inquiry_count']} total · {r['closed_order_count']} closed"

    # 4th element = relative PDF column width (see render_report_pdf's
    # width-weight comment). Contact now carries three merged fields so
    # it gets the biggest share; Type/Inquiries are short so they shrink.
    columns = [
        ("partner_name", "Partner", "str", 1.2),
        ("partner_type", "Type", "badge:partner_type", 0.7),
        ("contact", "Contact", "str", 1.8),
        ("inquiries_summary", "Inquiries", "str", 0.95),
        ("total_sales", "Package Sales", "money", 1.0),
        ("last_inquiry_at", "Last Inquiry", "datetime", 1.0),
        ("created_at", "On File Since", "datetime", 1.0),
    ]
    return columns, rows, len(rows) == MAX_ROWS, "Snapshot as of now"


def _report_packages(filters, branch_scope):
    """One row per package. Reference/order totals mirror package_detail's
    footer math (sum of qty * products.price, then the package's own
    discount_percent applied) — see packages.html / package_detail.html.
    """
    where, params = "", []
    if filters["package_scope"] != "all":
        where += " AND pk.partner_scope = %s"
        params.append(filters["package_scope"])
    if filters["package_status"] != "all":
        where += " AND pk.is_active = %s"
        params.append(filters["package_status"] == "Active")
    if filters["search"]:
        where += " AND (pk.package_name LIKE %s OR pk.description LIKE %s)"
        like = f"%{filters['search']}%"
        params += [like, like]

    rows = query(
        f"""SELECT pk.package_name, pk.description, pk.partner_scope, pk.discount_percent,
                   pk.is_active, pk.created_at,
                   COUNT(pki.package_item_id) AS item_count,
                   COALESCE(SUM(pki.qty * pr.price), 0) AS reference_total
            FROM packages pk
            LEFT JOIN package_items pki ON pki.package_id = pk.package_id
            LEFT JOIN products pr ON pki.sku = pr.sku
            WHERE 1=1 {where}
            GROUP BY pk.package_id
            ORDER BY pk.package_name LIMIT {MAX_ROWS}""",
        tuple(params),
    )
    for r in rows:
        r["description"] = r["description"] or "—"
        discount = float(r["discount_percent"])
        r["discount_percent"] = f"{discount:.2f}%"
        reference_total = float(r["reference_total"])
        r["reference_total"] = reference_total
        r["discounted_total"] = round(
            reference_total * (1 - discount / 100), 2)
        r["is_active"] = "Active" if r["is_active"] else "Retired"

    columns = [
        ("package_name", "Package", "str"), ("description", "Description", "str"),
        ("partner_scope", "Available To",
         "badge:package_scope"), ("item_count", "Items", "int"),
        ("reference_total", "Reference Value",
         "money"), ("discount_percent", "Discount", "str"),
        ("discounted_total", "Order Price",
         "money"), ("is_active", "Status", "badge:package_status"),
        ("created_at", "Created", "datetime"),
    ]
    return columns, rows, len(rows) == MAX_ROWS, "Snapshot as of now"


def _report_partner_inquiries(filters, branch_scope):
    """One row per inquiry — the same permanent, never-edited-except-
    status/remarks history shown on the Partner Inquiries page. Only
    Closed inquiries represent actual revenue (order_amount); everything
    else is still a lead — see schema.sql's partner_inquiries comment
    and partners.html/dashboard.html's matching footnotes.
    """
    where, params = "", []
    if filters["partner_type"] != "all":
        where += " AND pi.partner_type = %s"
        params.append(filters["partner_type"])
    if filters["inquiry_status"] != "all":
        where += " AND pi.status = %s"
        params.append(filters["inquiry_status"])
    if filters["search"]:
        where += (" AND (pi.company_name LIKE %s OR pi.contact_person LIKE %s "
                  "OR pi.package_name_snapshot LIKE %s OR pi.remarks LIKE %s)")
        like = f"%{filters['search']}%"
        params += [like, like, like, like]
    time_where, order, limit_n, truncated = _time_window(
        "pi.created_at", filters, params)
    where += time_where

    rows = query(
        f"""SELECT pi.created_at, pi.company_name, pi.partner_type, pi.contact_person, pi.phone,
                   pi.email, pi.package_name_snapshot, pi.order_amount, pi.status, pi.remarks
            FROM partner_inquiries pi
            WHERE 1=1 {where} {order} LIMIT {limit_n}""",
        tuple(params),
    )
    for r in rows:
        contact_person = r.pop("contact_person")
        r["company_name"] = (
            f"{r['company_name']} — {contact_person}" if contact_person else r["company_name"]
        )
        r["contact"] = " · ".join(p for p in (
            r.pop("phone"), r.pop("email")) if p) or "—"
        # Nullable — see schema.sql's comment: NULL means "unknown"
        # (older row predating this column), never coerced to 0.
        r["order_amount"] = float(
            r["order_amount"]) if r["order_amount"] is not None else None
        r["remarks"] = r["remarks"] or "—"

    # See _report_partners' comment above — same reasoning: Contact
    # Person now rides along inside Company, and Phone/Email collapse
    # into one Contact column, so both the column count and the short
    # badge/status columns shrink; Remarks/Package/Contact (the ones
    # that actually need room) grow. Address and HQ Notified (whether
    # the notification email happened to send) are dropped outright —
    # Address is almost always blank, and HQ Notified is an internal
    # mailer-ops flag, not information about the partner or the sale.
    columns = [
        ("created_at", "When", "datetime", 0.9),
        ("company_name", "Company", "str", 1.3),
        ("partner_type", "Type", "badge:partner_type", 0.7),
        ("contact", "Contact", "str", 1.5),
        ("package_name_snapshot", "Package", "str", 1.15),
        ("order_amount", "Order Amount (if Closed)", "money", 1.05),
        ("status", "Status", "badge:inquiry_status", 0.8),
        ("remarks", "Remarks (internal)", "str", 1.7),
    ]
    truncated = truncated and len(rows) == MAX_ROWS
    return columns, rows, truncated, _window_note(filters, truncated)


_BUILDERS = {
    "products": _report_products,
    "production_log": _report_production_log,
    "branch_stock": _report_branch_stock,
    "stock_requests": _report_stock_requests,
    "inventory_log": _report_inventory_log,
    "sales_history": _report_sales_history,
    "employee_purchases": _report_employee_purchases,
    "accounts": _report_accounts,
    "partners": _report_partners,
    "packages": _report_packages,
    "partner_inquiries": _report_partner_inquiries,
}


def get_report(report_type, filters, branch_scope=None, actor_label=""):
    """Build a report dict: title, subtitle, window_note, columns, rows, truncated.

    branch_scope: pass the signed-in branch's branch_id to force every
    query above to that branch and drop the Branch column entirely (used
    by routes/branch.py). Leave None for the admin side, where the
    branch_id filter (or "all branches") from the querystring applies.
    """
    if report_type not in _BUILDERS:
        raise ValueError(f"Unknown report type: {report_type}")

    meta = REPORT_TYPES[report_type]
    label = meta["branch_label"] if (
        branch_scope is not None and "branch_label" in meta) else meta["label"]

    columns, rows, truncated, window_note = _BUILDERS[report_type](
        filters, branch_scope)

    scoped_branch_name = _branch_name(branch_scope) if branch_scope is not None else (
        _branch_name(filters.get("branch_id")) if filters.get(
            "branch_id") not in (None, "all") else None
    )
    subtitle_bits = [scoped_branch_name] if scoped_branch_name else []
    if meta["windowed"]:
        subtitle_bits.append(window_note)
    subtitle = " · ".join(subtitle_bits) if subtitle_bits else window_note

    return {
        "report_type": report_type,
        "title": label,
        "subtitle": subtitle,
        "actor_label": actor_label,
        "generated_at": datetime.datetime.now(),
        "columns": columns,
        "rows": rows,
        "row_count": len(rows),
        "truncated": truncated,
    }


# ---------------------------------------------------------------- PDF rendering
def _pdf_styles():
    base = getSampleStyleSheet()
    return {
        "brand": ParagraphStyle("brand", parent=base["Normal"], fontName="DejaVuSans-Bold",
                                fontSize=16, textColor=INK, leading=19),
        "brand_sub": ParagraphStyle("brand_sub", parent=base["Normal"], fontName="DejaVuSans",
                                    fontSize=8, textColor=INK_FAINT, leading=11),
        "doc_title": ParagraphStyle("doc_title", parent=base["Normal"], fontName="DejaVuSans-Bold",
                                    fontSize=13, textColor=ACCENT_INK, alignment=TA_RIGHT, leading=16),
        "doc_meta": ParagraphStyle("doc_meta", parent=base["Normal"], fontName="DejaVuSans",
                                   fontSize=8.5, textColor=INK_FAINT, alignment=TA_RIGHT, leading=12),
        "th": ParagraphStyle("th", parent=base["Normal"], fontName="DejaVuSans-Bold",
                             fontSize=7.6, textColor=colors.white, leading=10),
        "td": ParagraphStyle("td", parent=base["Normal"], fontName="DejaVuSans",
                             fontSize=7.6, textColor=INK, leading=10),
        "td_num": ParagraphStyle("td_num", parent=base["Normal"], fontName="DejaVuSans",
                                 fontSize=7.6, textColor=INK, leading=10, alignment=TA_RIGHT),
        "footer": ParagraphStyle("footer", parent=base["Normal"], fontName="DejaVuSans",
                                 fontSize=7.3, textColor=INK_FAINT, leading=10),
    }


def _fmt_cell(value, ctype):
    if value is None:
        return "—"
    if ctype == "money":
        return f"₱{float(value):,.2f}"
    if ctype == "int":
        return f"{int(value):,}"
    if ctype == "datetime":
        return value.strftime("%b %d, %Y %I:%M %p") if isinstance(value, (datetime.date, datetime.datetime)) else str(value)
    return str(value)


def render_report_pdf(report):
    """Return a BytesIO PDF for a report dict built by get_report()."""
    s = _pdf_styles()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(letter),
        topMargin=16 * mm, bottomMargin=14 * mm, leftMargin=16 * mm, rightMargin=16 * mm,
        title=f"{report['title']} Report",
    )
    story = []

    header = Table(
        [[
            Table([[Paragraph(
                "<font color='#2E5AF0'>Heaven</font> <font color='#5B6272'>&amp;</font> "
                "<font color='#E23A48'>Angel</font> Scents", s["brand"])],
                [Paragraph("Perfume Manufacturing &amp; Retail &middot; Inventory System", s["brand_sub"])]],
                colWidths=[130 * mm]),
            Table([[Paragraph(report["title"].upper() + " REPORT", s["doc_title"])],
                   [Paragraph(report["subtitle"], s["doc_meta"])],
                   [Paragraph(
                       f"Generated {report['generated_at'].strftime('%b %d, %Y %I:%M %p')}"
                       + (f" by {report['actor_label']}" if report["actor_label"] else ""),
                       s["doc_meta"])]],
                  colWidths=[135 * mm]),
        ]],
        colWidths=[130 * mm, 135 * mm],
    )
    story.append(header)
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", thickness=1.2,
                 color=ACCENT, spaceAfter=10))

    # Columns are (key, label, ctype) from most report builders, or
    # (key, label, ctype, weight) from ones that specify custom relative
    # widths (see the width-weight comment further down) — normalize to
    # the 4-tuple shape once here so every loop below can unpack it the
    # same way regardless of which builder produced it.
    columns = [c if len(c) == 4 else (*c, 1) for c in report["columns"]]
    rows = report["rows"]

    if not rows:
        story.append(Spacer(1, 30))
        story.append(
            Paragraph("No data matches the selected filters.", s["footer"]))
    else:
        num_types = ("money", "int")
        header_row = [Paragraph(label, s["th"]) for _, label, _, _ in columns]
        data = [header_row]
        # Collect (row, col, bg_hex) for every badge-kind cell that
        # matched a known value, so the table style below can paint
        # just that cell — same colored-pill look as the web UI,
        # instead of every column rendering as flat black text.
        badge_cells = []
        for r_idx, row in enumerate(rows, start=1):
            cells = []
            for c_idx, (key, _, ctype, _) in enumerate(columns):
                value = row.get(key)
                if ctype.startswith("badge:"):
                    kind = ctype.split(":", 1)[1]
                    badge = _badge_colors(kind, value)
                    text = _fmt_cell(value, "str")
                    if badge:
                        bg_hex, text_hex = badge
                        badge_cells.append((r_idx, c_idx, bg_hex))
                        badge_style = ParagraphStyle(
                            f"badge_{r_idx}_{c_idx}", parent=s["td"],
                            textColor=colors.HexColor(text_hex), fontName="DejaVuSans-Bold",
                        )
                        cells.append(Paragraph(text, badge_style))
                    else:
                        cells.append(Paragraph(text, s["td"]))
                else:
                    cells.append(Paragraph(
                        _fmt_cell(value, ctype), s["td_num"] if ctype in num_types else s["td"]))
            data.append(cells)

        # Columns default to equal width (weight 1 each). A report can
        # give a column a different weight (see e.g. _report_partners /
        # _report_partner_inquiries) when an even split leaves some
        # columns too narrow to hold their content without ugly
        # mid-word wraps (a badge word breaking across two lines, a
        # phone number splitting mid-digit) while others sit mostly
        # empty — narrow columns shrink, columns that need the room
        # (Email, Address, Remarks, ...) grow to take up the slack.
        available_width = doc.width
        n = len(columns)
        weights = [w for _, _, _, w in columns]
        total_weight = sum(weights) or n
        col_widths = [available_width * (w / total_weight) for w in weights]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ]
        # Badge backgrounds are appended after ROWBACKGROUNDS so they
        # win for their specific cell — later commands override earlier
        # ones on the same cell in ReportLab's TableStyle.
        for r_idx, c_idx, bg_hex in badge_cells:
            style.append(
                ("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), colors.HexColor(bg_hex)))
        table.setStyle(TableStyle(style))
        story.append(table)

        story.append(Spacer(1, 8))
        count_note = f"{report['row_count']:,} row{'s' if report['row_count'] != 1 else ''} shown."
        if report["truncated"]:
            count_note += f" Results were capped at {MAX_ROWS:,} rows — narrow the filters for a complete report."
        story.append(Paragraph(count_note, s["footer"]))

    story.append(Spacer(1, 14))
    story.append(HRFlowable(width="100%", thickness=0.5,
                 color=BORDER, spaceAfter=6))
    story.append(Paragraph(
        "Generated from the Heaven &amp; Angel Scents inventory system. Figures reflect the underlying "
        "data at the moment this report was generated.",
        s["footer"],
    ))

    doc.build(story)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------- Excel rendering
def render_report_excel(report):
    """Return a BytesIO .xlsx for a report dict built by get_report()."""
    wb = Workbook()
    ws = wb.active
    ws.title = report["title"][:31] or "Report"

    # Some reports' columns carry a 4th "PDF width weight" element (see
    # render_report_pdf) — irrelevant here since Excel auto-sizes each
    # column from its own content below, so only the first three fields
    # are kept.
    columns = [c[:3] for c in report["columns"]]
    rows = report["rows"]
    n_cols = len(columns)

    # ---- title block ----
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=max(n_cols, 1))
    title_cell = ws.cell(row=1, column=1)
    # Rich text so "Heaven" / "Angel" render in the same two brand
    # colors as the sidebar, login page, and PDF report header,
    # instead of the whole title printing in one flat dark color.
    title_cell.value = CellRichText(
        TextBlock(InlineFont(rFont="Calibri", sz=14,
                  b=True, color="2E5AF0"), "Heaven "),
        TextBlock(InlineFont(rFont="Calibri", sz=14,
                  b=True, color="5B6272"), "& "),
        TextBlock(InlineFont(rFont="Calibri", sz=14,
                  b=True, color="E23A48"), "Angel "),
        TextBlock(InlineFont(rFont="Calibri", sz=14, b=True, color="12141A"),
                  f"Scents — {report['title']} Report"),
    )
    title_cell.fill = XL_TITLE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=max(n_cols, 1))
    meta = report["subtitle"]
    meta += f"  ·  Generated {report['generated_at'].strftime('%Y-%m-%d %H:%M')}"
    if report["actor_label"]:
        meta += f" by {report['actor_label']}"
    meta_cell = ws.cell(row=2, column=1, value=meta)
    meta_cell.font = Font(name="Calibri", size=9.5,
                          italic=True, color="5B6272")

    header_row_idx = 4
    if not rows:
        ws.cell(row=header_row_idx, column=1,
                value="No data matches the selected filters.").font = Font(italic=True)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ---- header ----
    for c, (_, label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row_idx, column=c, value=label)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = XL_HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        cell.border = XL_BORDER
    ws.row_dimensions[header_row_idx].height = 20

    # ---- data ----
    col_widths = [len(label) for _, label, _ in columns]
    for r_off, row in enumerate(rows):
        r = header_row_idx + 1 + r_off
        for c, (key, _, ctype) in enumerate(columns, start=1):
            value = row.get(key)
            cell = ws.cell(row=r, column=c)
            cell.border = XL_BORDER
            if value is None:
                cell.value = "—"
            elif ctype == "money":
                cell.value = float(value)
                cell.number_format = XL_MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            elif ctype == "int":
                cell.value = int(value)
                cell.alignment = Alignment(horizontal="right")
            elif ctype == "datetime" and isinstance(value, (datetime.date, datetime.datetime)):
                cell.value = value
                cell.number_format = XL_DATE_FMT
            elif ctype.startswith("badge:"):
                cell.value = str(value)
                # Same fill/text color as that value's badge on screen
                # (see BADGE_STYLES/BADGE_KIND_MAPS above) — falls back
                # to plain text if the value doesn't map to a badge.
                badge = _badge_colors(ctype.split(":", 1)[1], value)
                if badge:
                    bg_hex, text_hex = badge
                    cell.fill = PatternFill(
                        "solid", fgColor=bg_hex.lstrip("#"))
                    cell.font = Font(
                        name="Calibri", size=10, bold=True, color=text_hex.lstrip("#"))
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center")
            else:
                cell.value = str(value)
            width = len(str(cell.value)) if cell.value is not None else 0
            if width > col_widths[c - 1]:
                col_widths[c - 1] = width

    last_col_letter = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{header_row_idx}:{last_col_letter}{header_row_idx + len(rows)}"
    ws.freeze_panes = f"A{header_row_idx + 1}"

    for c, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(
            c)].width = min(max(width + 3, 10), 42)

    footer_row = header_row_idx + len(rows) + 2
    note = f"{report['row_count']:,} row(s) shown."
    if report["truncated"]:
        note += f" Results were capped at {MAX_ROWS:,} rows — narrow the filters for a complete report."
    ws.cell(row=footer_row, column=1, value=note).font = Font(
        size=9, italic=True, color="5B6272")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
